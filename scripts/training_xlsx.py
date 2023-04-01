from openpyxl import load_workbook
import csv

# Set the input and output file names
input_file = "./data/BCIT student program - COV 2022 Storefronts Inventory.xlsx"
output_file = "./data/converted_xlsx.csv"

# Load the input file and read in the data
workbook = load_workbook(filename=input_file)
sheet = workbook.active
rows = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    # Create a dictionary where the keys are the column names
    row_dict = {
        'OID_': row[0],
        ' id': row[1],
        'area': row[2],
        'address': row[3],
        'unit': row[4],
        'civic_number': row[5],
        'street': row[6],
        'address_above_door': row[7],
        'business_name': row[8],
        'retailgroup': row[9],
        'floor_area_sf': row[10],
        'year_recorded': row[11]
    }
    rows.append(row_dict)

# Filter the rows based on whether the floor_area_sf column has a value
filtered_rows = []
for row in rows:
    if row ['floor_area_sf']:
        if str(row['floor_area_sf']).isnumeric():
            filtered_rows.append(row)

# Open the output file and write the filtered rows
with open(output_file, 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=filtered_rows[0].keys())
    writer.writeheader()
    writer.writerows(filtered_rows)